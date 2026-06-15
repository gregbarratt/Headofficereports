from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.reporting import AuditLog, Booking, UploadBatch


COLUMN_ALIASES = {
    "status": ("status",),
    "customer_last_name": ("last name", "customer"),
    "agent_in_charge": ("agent in charge", "agent", "consultant", "consultant name"),
    "destination": ("destination",),
    "travel_elements_raw": ("elements",),
    "booking_ref": ("booking reference", "booking ref", "ref"),
    "departure_date": ("departure date", "depart"),
    "return_date": ("returned date", "return date"),
    "passenger_count": ("pax", "passengers", "passenger count"),
    "booking_date": ("date booked", "booking date", "booked"),
    "customer_balance_due_date": ("due date",),
    "imported_customer_outstanding": ("outstanding", "balance", "balance due"),
    "imported_supplier_outstanding": ("outstanding supplier", "outstanding supp"),
    "gross_booking_value": ("total cost", "gross", "total sell"),
    "expected_supplier_nett": ("nett", "net"),
    "non_trusted_total_received": ("total received", "paid", "total paid"),
    "non_trusted_paid_supplier": ("paid supp", "paid supplier"),
    "non_trusted_projected_profit": ("profit projected", "projected profit"),
}

MONEY_FIELDS = {
    "imported_customer_outstanding",
    "imported_supplier_outstanding",
    "gross_booking_value",
    "expected_supplier_nett",
    "non_trusted_total_received",
    "non_trusted_paid_supplier",
    "non_trusted_projected_profit",
}


@dataclass
class MasterBookingImportResult:
    row_count: int = 0
    accepted_rows: int = 0
    rejected_rows: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def error_summary(self) -> str | None:
        if not self.errors:
            return None
        preview = self.errors[:10]
        remaining = len(self.errors) - len(preview)
        suffix = f" Plus {remaining} more error(s)." if remaining > 0 else ""
        return " ".join(preview) + suffix


def normalise_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def read_tabular_rows(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return read_csv_rows(content)
    if extension == ".xlsx":
        return read_xlsx_rows(content)
    raise ValueError("Unsupported file type.")


def make_unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique_headers = []
    for header in headers:
        header_text = str(header or "").strip()
        key = normalise_header(header_text)
        seen[key] = seen.get(key, 0) + 1
        if seen[key] == 1:
            unique_headers.append(header_text)
        else:
            unique_headers.append(f"{header_text} {seen[key]}")
    return unique_headers


def read_csv_rows(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1252")

    reader = csv.reader(StringIO(text))
    raw_headers = next(reader, [])
    headers = make_unique_headers(raw_headers)
    rows = []
    for row in reader:
        row_data = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        if any(str(value or "").strip() for value in row_data.values()):
            rows.append(row_data)
    return headers, rows


def read_xlsx_rows(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = make_unique_headers([str(value or "") for value in next(rows, [])])
        output = []
        for row in rows:
            row_data = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            if any(value is not None and str(value).strip() for value in row_data.values()):
                output.append(row_data)
        return headers, output
    finally:
        workbook.close()


def build_column_map(headers: list[str]) -> dict[str, str]:
    normalised_to_original = {normalise_header(header): header for header in headers}
    column_map = {}
    for target_field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            original_header = normalised_to_original.get(alias)
            if original_header is not None:
                column_map[target_field] = original_header
                break
    return column_map


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalise_booking_ref(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None

    cleaned = text.strip().upper()
    compact = re.sub(r"[-_\s]+", "", cleaned)
    old_otc_match = re.fullmatch(r"OTC(1\d{6,})", compact)
    if old_otc_match:
        return f"OTC{old_otc_match.group(1)}"

    numeric_match = re.fullmatch(r"\d+(?:\.0+)?", cleaned)
    if numeric_match:
        number = int(Decimal(cleaned))
        if str(number).startswith("1") and len(str(number)) >= 7:
            return f"OTC{number}"
        return f"OTC-{number:05d}"

    otc_match = re.fullmatch(r"OTC[-_\s]?0*(\d+)", cleaned)
    if otc_match:
        digits = otc_match.group(1)
        if digits.startswith("1") and len(digits) >= 7:
            return f"OTC{digits}"
        return f"OTC-{int(digits):05d}"

    return cleaned


def should_skip_master_row(row: dict[str, Any], column_map: dict[str, str]) -> bool:
    raw_ref = clean_text(get_row_value(row, column_map, "booking_ref"))
    if not raw_ref:
        return False

    normalised_ref = normalise_header(raw_ref)
    if normalised_ref in {"keep", "summary breakdown"}:
        return True
    if normalised_ref.startswith(("this should", "total ", "average gross")):
        return True
    return False


def parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace(",", "")
    try:
        return int(Decimal(cleaned))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid whole number value '{value}'.") from exc


def parse_money(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int | float):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = str(value).strip()
    if not text or text in {"-", "--"}:
        return None

    is_negative = text.startswith("(") and text.endswith(")")
    cleaned = (
        text.replace("£", "")
        .replace("Ł", "")
        .replace("GBP", "")
        .replace("?", "")
        .replace(",", "")
        .replace("(", "")
        .replace(")", "")
        .replace(" ", "")
    )
    if cleaned in {"", "-"}:
        return None

    try:
        amount = Decimal(cleaned).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid money value '{value}'.") from exc

    return -amount if is_negative else amount


def parse_date(value: Any) -> date | None:
    parsed = parse_datetime(value)
    return parsed.date() if parsed else None


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = str(value).strip()
    if not text:
        return None

    formats = (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    )
    for date_format in formats:
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    raise ValueError(f"Invalid date value '{value}'.")


def normalise_booking_status(status: str | None) -> str | None:
    cleaned = (status or "").strip().lower()
    if not cleaned:
        return None
    status_map = {
        "chg": "amended/live",
        "changed": "amended/live",
        "amended": "amended/live",
        "complete": "completed",
        "completed": "completed",
        "can": "cancelled",
        "cancelled": "cancelled",
        "canceled": "cancelled",
        "open": "open",
        "live": "open",
    }
    return status_map.get(cleaned, cleaned)


def determine_booking_company(booking_ref: str | None) -> str:
    reference = (booking_ref or "").strip().upper()
    if reference.startswith("OTC"):
        return "otc"
    if reference.startswith(("LEM", "LMX", "LM-", "LM_")) or "LEMIEUX" in reference:
        return "lemieux"
    return "review"


def parse_elements(elements: str | None) -> dict[str, bool]:
    cleaned = (elements or "").lower()
    return {
        "flight_included": any(term in cleaned for term in ("flight", "flights", "airline", "flt")),
        "accommodation_included": any(term in cleaned for term in ("accommodation", "accom", "hotel", "htl")),
        "cruise_included": "cruise" in cleaned,
        "extras_included": any(
            term in cleaned
            for term in ("extra", "extras", "transfer", "tran", "car hire", "hire", "insurance", "ticket", "tkt", "att")
        ),
        "package_included": any(term in cleaned for term in ("package", "packages", "multipackage")),
    }


def determine_atol_review_status(flags: dict[str, bool]) -> str:
    has_flight = flags["flight_included"]
    has_accommodation = flags["accommodation_included"]
    has_cruise = flags["cruise_included"]
    has_extras = flags["extras_included"]
    has_package = flags["package_included"]

    if has_flight and has_accommodation:
        return "ATOL required"
    if has_flight and has_cruise:
        return "ATOL review"
    if has_flight and (has_extras or has_package):
        return "ATOL review / likely required"
    if has_flight:
        return "flight-only review"
    if has_package:
        return "ATOL review"
    return "non-flight / non-ATOL review"


def get_row_value(row: dict[str, Any], column_map: dict[str, str], field_name: str) -> Any:
    header = column_map.get(field_name)
    if header is None:
        return None
    return row.get(header)


def import_master_booking_report(
    db: Session,
    upload_batch: UploadBatch,
    filename: str,
    content: bytes,
    actor_user_id: int | None,
) -> MasterBookingImportResult:
    headers, rows = read_tabular_rows(filename, content)
    result = MasterBookingImportResult()
    column_map = build_column_map(headers)

    if "booking_ref" not in column_map:
        result.rejected_rows = len(rows)
        result.row_count = len(rows)
        result.errors.append("Booking Reference column is required for master booking imports.")
        return result

    for index, row in enumerate(rows, start=2):
        if should_skip_master_row(row, column_map):
            continue

        result.row_count += 1

        try:
            booking_ref = normalise_booking_ref(get_row_value(row, column_map, "booking_ref"))
            if not booking_ref:
                raise ValueError("Booking Reference is missing.")

            imported_status = clean_text(get_row_value(row, column_map, "status"))
            elements_raw = clean_text(get_row_value(row, column_map, "travel_elements_raw"))
            flags = parse_elements(elements_raw)

            values = {
                "booking_ref": booking_ref,
                "booking_company": determine_booking_company(booking_ref),
                "imported_booking_status": imported_status,
                "normalised_status": normalise_booking_status(imported_status),
                "customer_last_name": clean_text(get_row_value(row, column_map, "customer_last_name")),
                "agent_in_charge": clean_text(get_row_value(row, column_map, "agent_in_charge")),
                "destination": clean_text(get_row_value(row, column_map, "destination")),
                "travel_elements_raw": elements_raw,
                "departure_date": parse_date(get_row_value(row, column_map, "departure_date")),
                "return_date": parse_date(get_row_value(row, column_map, "return_date")),
                "passenger_count": parse_int(get_row_value(row, column_map, "passenger_count")),
                "booking_date": parse_datetime(get_row_value(row, column_map, "booking_date")),
                "customer_balance_due_date": parse_date(get_row_value(row, column_map, "customer_balance_due_date")),
                "last_master_upload_batch_id": upload_batch.id,
                **flags,
            }

            for money_field in MONEY_FIELDS:
                values[money_field] = parse_money(get_row_value(row, column_map, money_field))

            values["atol_review_status"] = determine_atol_review_status(flags)

            booking = db.scalar(select(Booking).where(Booking.booking_ref == booking_ref))
            if booking is None:
                booking = Booking(**values)
                db.add(booking)
            else:
                for field_name, value in values.items():
                    setattr(booking, field_name, value)

            result.accepted_rows += 1
        except ValueError as exc:
            result.rejected_rows += 1
            result.errors.append(f"Row {index}: {exc}")

    db.add(
        AuditLog(
            actor_user_id=actor_user_id,
            action="master_booking_import",
            table_name="upload_batches",
            record_id=upload_batch.id,
            description=(
                f"Master booking import accepted {result.accepted_rows} row(s) "
                f"and rejected {result.rejected_rows} row(s)."
            ),
        )
    )
    return result

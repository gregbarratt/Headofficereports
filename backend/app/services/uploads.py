from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4

from fastapi import UploadFile
from openpyxl import load_workbook

from app.core.config import settings
from app.core.uploads import ALLOWED_UPLOAD_EXTENSIONS


@dataclass
class FileValidationResult:
    row_count: int
    accepted_rows: int
    rejected_rows: int
    error_summary: str | None = None

    @property
    def passed(self) -> bool:
        return self.error_summary is None


def clean_filename(filename: str) -> str:
    return Path(filename or "upload").name.replace("/", "_").replace("\\", "_")


def build_stored_filename(batch_id: int, original_filename: str) -> str:
    extension = Path(original_filename).suffix.lower()
    return f"upload_batch_{batch_id}_{uuid4().hex}{extension}"


def validate_extension(filename: str) -> str | None:
    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_UPLOAD_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_UPLOAD_EXTENSIONS))
        return f"Unsupported file type. Please upload one of: {allowed}."
    return None


def validate_file_size(content: bytes) -> str | None:
    if len(content) > settings.max_upload_size_bytes:
        return f"File is too large. Maximum allowed size is {settings.max_upload_size_mb}MB."
    return None


def inspect_upload_content(filename: str, content: bytes) -> FileValidationResult:
    extension_error = validate_extension(filename)
    if extension_error:
        return FileValidationResult(0, 0, 0, extension_error)

    size_error = validate_file_size(content)
    if size_error:
        return FileValidationResult(0, 0, 0, size_error)

    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return inspect_csv_content(content)
    if extension == ".xlsx":
        return inspect_xlsx_content(content)

    return FileValidationResult(0, 0, 0, "Unsupported file type.")


def inspect_csv_content(content: bytes) -> FileValidationResult:
    try:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1252")

        reader = csv.reader(StringIO(text))
        rows = list(reader)
    except csv.Error as exc:
        return FileValidationResult(0, 0, 0, f"CSV could not be read: {exc}")

    if not rows:
        return FileValidationResult(0, 0, 0, "File is empty.")

    header = rows[0]
    if not any(str(cell).strip() for cell in header):
        return FileValidationResult(0, 0, 0, "Header row is missing.")

    data_rows = [row for row in rows[1:] if any(str(cell).strip() for cell in row)]
    row_count = len(data_rows)
    return FileValidationResult(row_count=row_count, accepted_rows=row_count, rejected_rows=0)


def inspect_xlsx_content(content: bytes) -> FileValidationResult:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None or not any(str(cell).strip() for cell in header if cell is not None):
            return FileValidationResult(0, 0, 0, "Header row is missing.")

        row_count = 0
        for row in rows:
            if any(cell is not None and str(cell).strip() for cell in row):
                row_count += 1
    except Exception as exc:
        return FileValidationResult(0, 0, 0, f"XLSX could not be read: {exc}")
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return FileValidationResult(row_count=row_count, accepted_rows=row_count, rejected_rows=0)


async def read_upload_file(upload_file: UploadFile) -> bytes:
    return await upload_file.read()


def save_upload_file(content: bytes, stored_filename: str) -> Path:
    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    target_path = settings.upload_dir / stored_filename
    target_path.write_bytes(content)
    return target_path

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.reporting import (
    AgentCommission,
    Booking,
    CustomerPayment,
    ExceptionRecord,
    InsuranceCost,
    Refund,
    ReportRun,
    SupplierPayment,
    WeeklySnapshot,
)
from app.services.insurance_import import is_active_insurance_status
from app.services.trust_reconciliation import calculate_trust_reconciliation
from app.services.weekly_snapshots import compare_snapshots, previous_snapshot, snapshot_rows


ZERO = Decimal("0.00")

REPORT_TYPES = {
    "executive_weekly_overview": "Executive Weekly Overview",
    "trust_reconciliation": "Trust Reconciliation Report",
    "customer_payments": "Customer Payments Report",
    "supplier_payments": "Supplier Payments Report",
    "supplier_liability": "Supplier Liability Report",
    "insurance_costs": "Insurance Costs Report",
    "refund_liability": "Refund Liability Report",
    "agent_commission": "Agent Commission Report",
    "true_booking_profitability": "True Booking Profitability Report",
    "atol_compliance": "ATOL Compliance Report",
    "week_on_week_movement": "Week-on-Week Movement Report",
    "exception_report": "Exception Report",
}


def money(value: Decimal | None) -> Decimal:
    if value is None:
        return ZERO
    return Decimal(value).quantize(Decimal("0.01"))


def positive(value: Decimal) -> Decimal:
    return value if value > ZERO else ZERO


def value_for_excel(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(money(value))
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    return value


def append_rows(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(sheet_name[:31])
    sheet.append(headers)
    for row in rows:
        sheet.append([value_for_excel(value) for value in row])

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)


def workbook_bytes(workbook: Workbook) -> bytes:
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def report_filename(report_type: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{report_type}_{timestamp}.xlsx"


def booking_counts(db: Session) -> dict[str, int]:
    rows = db.execute(select(Booking.normalised_status, func.count()).group_by(Booking.normalised_status))
    return {(status or "unknown"): count for status, count in rows}


def latest_snapshot(db: Session) -> WeeklySnapshot | None:
    return db.scalar(
        select(WeeklySnapshot).order_by(WeeklySnapshot.week_start_date.desc(), WeeklySnapshot.id.desc()).limit(1)
    )


def supplier_totals(db: Session) -> dict[str, Decimal]:
    totals: dict[str, Decimal] = {}
    for booking_ref, total in db.execute(
        select(SupplierPayment.booking_ref, func.sum(SupplierPayment.supplier_payment_amount))
        .where(SupplierPayment.booking_ref.is_not(None))
        .where(SupplierPayment.payment_source == "taps")
        .group_by(SupplierPayment.booking_ref)
    ):
        totals[booking_ref] = money(total)
    return totals


def insurance_totals(db: Session) -> dict[str, Decimal]:
    totals: dict[str, Decimal] = {}
    for cost in db.scalars(select(InsuranceCost)):
        if cost.booking_ref and is_active_insurance_status(cost.insurance_status):
            totals[cost.booking_ref] = money(totals.get(cost.booking_ref, ZERO) + money(cost.insurance_cost_amount))
    return totals


def refund_unpaid(refund: Refund) -> Decimal:
    return positive(money(refund.refund_amount_due) - money(refund.refund_amount_paid))


def build_executive_weekly_overview(db: Session, workbook: Workbook) -> None:
    trust = calculate_trust_reconciliation(db)
    counts = booking_counts(db)
    snapshot = latest_snapshot(db)
    open_exceptions = db.scalar(
        select(func.count()).select_from(ExceptionRecord).where(ExceptionRecord.status.in_(("open", "reviewing")))
    )
    critical_exceptions = db.scalar(
        select(func.count()).select_from(ExceptionRecord).where(ExceptionRecord.severity == "critical")
    )

    rows = [
        ["Actual Trust Balance", trust.summary.actual_trust_balance],
        ["Required Trust Balance", trust.summary.required_trust_balance],
        ["Trust Variance", trust.summary.trust_variance],
        ["Live/Open Bookings", counts.get("open", 0) + counts.get("amended/live", 0)],
        ["Completed Bookings", counts.get("completed", 0)],
        ["Cancelled Bookings", counts.get("cancelled", 0)],
        ["Refunds Due", trust.summary.refunds_due],
        ["Supplier Payments Made", trust.summary.supplier_payments_made],
        ["Open/Reviewing Exceptions", open_exceptions or 0],
        ["Critical Exceptions", critical_exceptions or 0],
        ["Latest Snapshot Week", f"{snapshot.week_start_date} to {snapshot.week_end_date}" if snapshot else None],
    ]
    append_rows(workbook, "Executive Overview", ["Metric", "Value"], rows)


def build_trust_reconciliation(db: Session, workbook: Workbook) -> None:
    trust = calculate_trust_reconciliation(db)
    append_rows(
        workbook,
        "Trust Summary",
        ["Metric", "Value"],
        [
            ["Customer payments received", trust.summary.customer_payments_received],
            ["Card fees", trust.summary.card_fees],
            ["Net trust receipts", trust.summary.net_trust_receipts],
            ["Supplier payments made", trust.summary.supplier_payments_made],
            ["Refunds paid", trust.summary.refunds_paid],
            ["Refunds due", trust.summary.refunds_due],
            ["Required trust balance", trust.summary.required_trust_balance],
            ["Actual trust balance", trust.summary.actual_trust_balance],
            ["Trust variance", trust.summary.trust_variance],
        ],
    )
    append_rows(
        workbook,
        "Booking Trust",
        [
            "Booking Ref",
            "Last Name",
            "Status",
            "Gross Value",
            "Customer Paid",
            "Card Fees",
            "Net Trust Receipts",
            "Supplier Paid",
            "Refunds Paid",
            "Refunds Unpaid",
            "Current Trust Balance",
            "Required Contribution",
            "Trust Status",
            "Missing Data",
        ],
        [
            [
                row.booking_ref,
                row.customer_last_name,
                row.booking_status,
                row.gross_booking_value,
                row.customer_payments_received,
                row.card_fees,
                row.net_trust_receipts,
                row.supplier_payments_made,
                row.refunds_paid,
                row.refunds_unpaid,
                row.current_booking_trust_balance,
                row.required_trust_balance_contribution,
                row.trust_status,
                "; ".join(row.missing_items),
            ]
            for row in trust.bookings
        ],
    )


def build_customer_payments(db: Session, workbook: Workbook) -> None:
    payments = list(db.scalars(select(CustomerPayment).order_by(CustomerPayment.created_at.desc())))
    append_rows(
        workbook,
        "Customer Payments",
        [
            "Transaction ID",
            "Source",
            "Booking Ref",
            "Invoice Ref",
            "Customer Name",
            "Payment Date",
            "Settlement Date",
            "Gross Amount",
            "Fee Amount",
            "Net Settled",
            "Fee Estimated",
            "Payment Method",
            "Card Type",
            "Card Brand",
            "Status",
            "Refund",
            "Chargeback",
            "Merchant Account",
            "Settlement Batch",
            "Match Confidence",
        ],
        [
            [
                payment.transaction_id,
                payment.payment_source,
                payment.booking_ref,
                payment.invoice_reference,
                payment.customer_name,
                payment.payment_date,
                payment.settlement_date,
                payment.gross_amount,
                payment.fee_amount,
                payment.net_settled_amount,
                payment.fee_is_estimated,
                payment.payment_method,
                payment.card_type,
                payment.card_brand,
                payment.transaction_status,
                payment.refund_indicator,
                payment.chargeback_indicator,
                payment.merchant_account,
                payment.settlement_batch_reference,
                payment.match_confidence,
            ]
            for payment in payments
        ],
    )


def build_supplier_payments(db: Session, workbook: Workbook) -> None:
    payments = list(db.scalars(select(SupplierPayment).order_by(SupplierPayment.created_at.desc())))
    append_rows(
        workbook,
        "Supplier Payments",
        [
            "Booking Ref",
            "Source",
            "Transaction Date",
            "Product",
            "Supplier",
            "Payment Supplier",
            "Payment Method",
            "Payment Value",
            "Associated VAT",
            "Duplicate",
            "Match Status",
        ],
        [
            [
                payment.booking_ref,
                payment.payment_source,
                payment.supplier_payment_date,
                payment.product_type,
                payment.supplier_name,
                payment.payment_supplier_name,
                payment.supplier_payment_method,
                payment.supplier_payment_amount,
                payment.associated_vat,
                payment.is_duplicate,
                payment.match_status,
            ]
            for payment in payments
        ],
    )


def build_supplier_liability(db: Session, workbook: Workbook) -> None:
    totals = supplier_totals(db)
    insurance_by_booking = insurance_totals(db)
    bookings = list(db.scalars(select(Booking).order_by(Booking.booking_ref)))
    rows = []
    for booking in bookings:
        supplier_nett = money(booking.expected_supplier_nett)
        insurance_cost = insurance_by_booking.get(booking.booking_ref, ZERO)
        expected = money(supplier_nett + insurance_cost)
        paid = totals.get(booking.booking_ref, ZERO)
        balance = money(expected - paid)
        if expected == ZERO and paid == ZERO:
            status = "no_supplier_nett"
        elif balance == ZERO:
            status = "paid_in_full"
        elif balance < ZERO:
            status = "overpaid"
        elif paid > ZERO:
            status = "partially_paid"
        else:
            status = "unpaid"
        rows.append([booking.booking_ref, booking.customer_last_name, supplier_nett, insurance_cost, expected, paid, balance, status])
    append_rows(
        workbook,
        "Supplier Liability",
        [
            "Booking Ref",
            "Last Name",
            "Expected Supplier Nett",
            "Insurance Cost",
            "Total Expected Booking Cost",
            "Supplier Paid",
            "Balance Due",
            "Status",
        ],
        rows,
    )


def build_insurance_costs(db: Session, workbook: Workbook) -> None:
    costs = list(db.scalars(select(InsuranceCost).order_by(InsuranceCost.created_at.desc())))
    append_rows(
        workbook,
        "Insurance Costs",
        [
            "Booking Ref",
            "External Reference",
            "Trade Code",
            "Trading Name",
            "Lead Name",
            "Departure Date",
            "Supplement Type",
            "Gross",
            "Discount",
            "Net",
            "Insurance Cost",
            "Status",
            "Match Status",
            "Duplicate",
        ],
        [
            [
                cost.booking_ref,
                cost.external_reference,
                cost.trade_code,
                cost.trading_name,
                cost.lead_name,
                cost.departure_date,
                cost.supplement_type,
                cost.gross_amount,
                cost.discount_amount,
                cost.net_amount,
                cost.insurance_cost_amount,
                cost.insurance_status,
                cost.match_status,
                cost.is_duplicate,
            ]
            for cost in costs
        ],
    )


def build_refund_liability(db: Session, workbook: Workbook) -> None:
    refunds = list(db.scalars(select(Refund).order_by(Refund.created_at.desc())))
    append_rows(
        workbook,
        "Refund Liability",
        [
            "Booking Ref",
            "Customer Name",
            "Reason",
            "Amount Due",
            "Amount Paid",
            "Unpaid",
            "Status",
            "Supplier Refund Expected",
            "Supplier Refund Received",
            "Due Date",
            "Paid Date",
        ],
        [
            [
                refund.booking_ref,
                refund.customer_name,
                refund.refund_reason,
                refund.refund_amount_due,
                refund.refund_amount_paid,
                refund_unpaid(refund),
                refund.refund_status,
                refund.supplier_refund_expected,
                refund.supplier_refund_received,
                refund.due_date,
                refund.paid_date,
            ]
            for refund in refunds
        ],
    )


def build_agent_commission(db: Session, workbook: Workbook) -> None:
    commissions = list(db.scalars(select(AgentCommission).order_by(AgentCommission.created_at.desc())))
    append_rows(
        workbook,
        "Agent Commission",
        [
            "Booking Ref",
            "Agent Name",
            "Commission Basis",
            "Gross Commission",
            "Deductions",
            "Net Commission Due",
            "Status",
            "Due Date",
            "Paid Date",
        ],
        [
            [
                commission.booking_ref,
                commission.agent_name,
                commission.commission_basis,
                commission.gross_commission,
                commission.deductions,
                commission.net_commission_due,
                commission.commission_status,
                commission.due_date,
                commission.paid_date,
            ]
            for commission in commissions
        ],
    )


def build_true_booking_profitability(db: Session, workbook: Workbook) -> None:
    bookings = list(db.scalars(select(Booking).order_by(Booking.booking_ref)))
    payments_by_booking: dict[str, list[CustomerPayment]] = {}
    commissions_by_booking: dict[str, list[AgentCommission]] = {}
    refunds_by_booking: dict[str, list[Refund]] = {}
    insurance_by_booking = insurance_totals(db)

    for payment in db.scalars(select(CustomerPayment).where(CustomerPayment.payment_source == "sings")):
        if payment.booking_ref and payment.match_confidence != "unmatched":
            payments_by_booking.setdefault(payment.booking_ref, []).append(payment)
    for commission in db.scalars(select(AgentCommission)):
        if commission.booking_ref:
            commissions_by_booking.setdefault(commission.booking_ref, []).append(commission)
    for refund in db.scalars(select(Refund)):
        if refund.booking_ref:
            refunds_by_booking.setdefault(refund.booking_ref, []).append(refund)

    rows = []
    for booking in bookings:
        card_fees = sum((money(payment.fee_amount) for payment in payments_by_booking.get(booking.booking_ref, [])), ZERO)
        commission_due = sum(
            (money(commission.net_commission_due) for commission in commissions_by_booking.get(booking.booking_ref, [])),
            ZERO,
        )
        refund_adjustments = sum(
            (money(refund.refund_amount_due) for refund in refunds_by_booking.get(booking.booking_ref, [])),
            ZERO,
        )
        insurance_cost = insurance_by_booking.get(booking.booking_ref, ZERO)
        true_profit = None
        margin = None
        if booking.gross_booking_value is not None and booking.expected_supplier_nett is not None:
            true_profit = (
                money(booking.gross_booking_value)
                - money(booking.expected_supplier_nett)
                - insurance_cost
                - card_fees
                - commission_due
                - refund_adjustments
            )
            if money(booking.gross_booking_value) != ZERO:
                margin = (true_profit / money(booking.gross_booking_value) * Decimal("100")).quantize(Decimal("0.01"))
        rows.append(
            [
                booking.booking_ref,
                booking.customer_last_name,
                booking.gross_booking_value,
                booking.expected_supplier_nett,
                insurance_cost,
                card_fees,
                commission_due,
                refund_adjustments,
                true_profit,
                margin,
            ]
        )
    append_rows(
        workbook,
        "True Profitability",
        [
            "Booking Ref",
            "Last Name",
            "Gross Booking Value",
            "Expected Supplier Nett",
            "Insurance Cost",
            "Card Fees",
            "Agent Commission",
            "Refunds / Adjustments",
            "True Profit",
            "True Margin %",
        ],
        rows,
    )


def build_atol_compliance(db: Session, workbook: Workbook) -> None:
    bookings = list(db.scalars(select(Booking).order_by(Booking.booking_ref)))
    append_rows(
        workbook,
        "ATOL Compliance",
        [
            "Booking Ref",
            "Last Name",
            "Destination",
            "Elements",
            "Flight",
            "Accommodation",
            "Cruise",
            "Extras",
            "Package",
            "ATOL Status",
            "Certificate Issued",
        ],
        [
            [
                booking.booking_ref,
                booking.customer_last_name,
                booking.destination,
                booking.travel_elements_raw,
                booking.flight_included,
                booking.accommodation_included,
                booking.cruise_included,
                booking.extras_included,
                booking.package_included,
                booking.atol_review_status,
                booking.atol_certificate_issued,
            ]
            for booking in bookings
        ],
    )


def build_week_on_week_movement(db: Session, workbook: Workbook) -> None:
    snapshot = latest_snapshot(db)
    if snapshot is None:
        append_rows(workbook, "Week Movement", ["Message"], [["No weekly snapshot has been generated yet."]])
        return
    previous = previous_snapshot(db, snapshot)
    if previous is None:
        append_rows(workbook, "Week Movement", ["Message"], [["No previous snapshot is available yet."]])
        return
    movements = compare_snapshots(snapshot_rows(db, previous.id), snapshot_rows(db, snapshot.id))
    append_rows(
        workbook,
        "Week Movement",
        ["Booking Ref", "Movement", "Field", "Previous", "Current", "Detail"],
        [
            [
                movement.booking_ref,
                movement.movement_type,
                movement.field_name,
                movement.previous_value,
                movement.current_value,
                movement.description,
            ]
            for movement in movements
        ],
    )


def build_exception_report(db: Session, workbook: Workbook) -> None:
    exceptions = list(db.scalars(select(ExceptionRecord).order_by(ExceptionRecord.detected_at.desc())))
    append_rows(
        workbook,
        "Exceptions",
        ["Type", "Severity", "Status", "Title", "Booking Ref", "Detail", "Related Table", "Related ID", "Detected", "Resolved"],
        [
            [
                exception.exception_type,
                exception.severity,
                exception.status,
                exception.title,
                exception.booking_ref,
                exception.detail,
                exception.related_table,
                exception.related_record_id,
                exception.detected_at,
                exception.resolved_at,
            ]
            for exception in exceptions
        ],
    )


REPORT_BUILDERS = {
    "executive_weekly_overview": build_executive_weekly_overview,
    "trust_reconciliation": build_trust_reconciliation,
    "customer_payments": build_customer_payments,
    "supplier_payments": build_supplier_payments,
    "supplier_liability": build_supplier_liability,
    "insurance_costs": build_insurance_costs,
    "refund_liability": build_refund_liability,
    "agent_commission": build_agent_commission,
    "true_booking_profitability": build_true_booking_profitability,
    "atol_compliance": build_atol_compliance,
    "week_on_week_movement": build_week_on_week_movement,
    "exception_report": build_exception_report,
}


def generate_excel_report(db: Session, report_type: str) -> tuple[str, bytes]:
    if report_type not in REPORT_BUILDERS:
        raise ValueError("Unknown report type.")

    workbook = Workbook()
    REPORT_BUILDERS[report_type](db, workbook)
    return report_filename(report_type), workbook_bytes(workbook)


def create_report_run(db: Session, report_type: str, actor_user_id: int | None) -> ReportRun:
    report_run = ReportRun(
        report_type=report_type,
        status="running",
        requested_by_user_id=actor_user_id,
    )
    db.add(report_run)
    db.flush()
    return report_run
